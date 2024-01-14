import openpyxl
import os


def export_table(header: list, data: list[list], title: str='Table', filepath: str='daten.xlsx', override: bool=True):
    """
    Export data in a table format to an Excel file.

    Args:
        header (list): A list containing column headers for the table.
        data (list[list]): A list of lists, where each inner list represents a row of data for the table.
        title (str, optional): The title of the Excel sheet. Default is 'Table'.
        filepath (str, optional): The filepath where the Excel file should be saved. Default is 'daten.xlsx'.
        override (bool, optional): If True, overwrite the file if it already exists. If False, return False and do not override.

    Returns:
        None or bool: None if the file was saved successfully, or False if the user chose not to override an existing file.

    Raises:
        PermissionError: If there is a permission issue while saving the file.
    """
    if not override and os.path.isfile(filepath):
        return False
    
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook['Sheet']
        worksheet.title = title

        # Populate column headers in the first row
        for col_num, header_text in enumerate(header, start=1):
            worksheet.cell(row=1, column=col_num, value=header_text)

        # Populate data rows in subsequent rows
        for row_num, row_data in enumerate(data, start=2):
            for col_num, cell_data in enumerate(row_data, start=1):
                worksheet.cell(row=row_num, column=col_num, value=cell_data)

        # Save the workbook to the specified filepath
        workbook.save(filepath)
        print(f"Data exported successfully to '{filepath}'")
    except PermissionError:
        print(f"Permission error: Cannot save to '{filepath}'. Please check file permissions.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    finally:
        workbook.close()  # Close the workbook

# Example usage:
header = ["Name", "Age", "City"]
data = [["Alice", 30, "New York"], ["Bob", 25, "Los Angeles"], ["Charlie", 35, "Chicago"]]

export_table(header, data, title="Person Data", filepath="person_data.xlsx")
